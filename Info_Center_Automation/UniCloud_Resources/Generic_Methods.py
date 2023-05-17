#------------------------------------------------------------------
#	Description			:		This file contains all the Generic methods.
#	Project				:		UniCloud
#	Author				:		Vikram Nagesh
#	C 2018 Systech International. All rights reserved
#------------------------------------------------------------------
#
#	Prologue:
#	- Not Applicable
#
#	Epilogue:
#	- Not Applicable

"""
List of Methods available:
	01. read_excel_file_and_return_worksheet_names(filepathandfilename)
	02. read_excel_file_and_return_worksheets_count(filepathandfilename)
	03. total_rows_from_worksheet(filepathandfilename, worksheetname)
	04. extract_info_center_file_name(stringValue)
	05. update_one_off_status_time_to_excel(filePathAndFileName, worksheetName, ColumnAlphabet)
	06. update_status_to_excel(filePathAndFileName, worksheetName, rowNumber, status, message, ColumnAlphabet):
"""

from openpyxl import load_workbook
import datetime

def read_excel_file_and_return_worksheet_names(filepathandfilename):
	wb = load_workbook(filepathandfilename)
	return wb.sheetnames
	
def read_excel_file_and_return_worksheets_count(filepathandfilename):
	wb = load_workbook(filepathandfilename)
	return wb.worksheets.__len__()

def total_rows_from_worksheet(filepathandfilename, worksheetname):
	wb = load_workbook(filepathandfilename)
	ws = wb[worksheetname]
	return ws.max_row
	
def extract_info_center_file_name(stringValue):
	if stringValue[-4]=='.':
		return str(stringValue).split('-')[0] + stringValue[-4:]
	elif stringValue[-5]=='.':
		return str(stringValue).split('-')[0] + stringValue[-5:]
	
def update_one_off_status_time_to_excel(filePathAndFileName, worksheetName, ColumnAlphabet):
	wb = load_workbook(filePathAndFileName, data_only=True)
	ws = wb[worksheetName]
	ws[ColumnAlphabet + '1'] = 'Status - ' + str(datetime.datetime.now())
	wb.close()
	wb.save(filePathAndFileName)
	
def update_status_to_excel(filePathAndFileName, worksheetName, rowNumber, status, message, ColumnAlphabet):
	wb = load_workbook(filePathAndFileName, data_only=True)
	ws = wb[worksheetName]
	ws[ColumnAlphabet + str(rowNumber)] = status + ' - ' + message
	wb.close()
	wb.save(filePathAndFileName)

def extract_info_center_file_name(stringValue):
	if stringValue[-4]=='.':
		return str(stringValue).split('-')[0] + stringValue[-4:]
	elif stringValue[-5]=='.':
		return str(stringValue).split('-')[0] + stringValue[-5:]

def getTaglengthgt15(tag1, tag2):
	if len(tag1) > 15 or len(tag2) > 15:
		return True
	else:
		return False
	
def getTaglengthgt15FromOneTagColumn(tag1):
	for individualTag in str(tag1).split(","):
		if len(individualTag) > 15:
			tagStatus = True
			break
		else:
			tagStatus = False
	return tagStatus